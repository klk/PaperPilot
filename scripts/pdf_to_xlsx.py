#!/usr/bin/env python3
"""Rebuild PDF tables as editable Excel cells and preserve prose by page."""

from __future__ import annotations

import argparse
import math
import re
from collections import defaultdict
from pathlib import Path

import fitz
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


THIN_GREY = Side(style="thin", color="BDC3C9")
FONT_NAME = "PingFang SC"


def unique(values: list[float], tolerance: float = 1.5) -> list[float]:
    output: list[float] = []
    for value in sorted(values):
        if not output or abs(value - output[-1]) > tolerance:
            output.append(value)
    return output


def rgb(fill) -> str:
    if not fill or len(fill) < 3:
        return "FFFFFF"
    return "".join(f"{max(0, min(255, round(value * 255))):02X}" for value in fill[:3])


def text_lines(page: fitz.Page):
    lines = []
    for block in page.get_text("dict").get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            text = "".join(span.get("text", "") for span in line.get("spans", [])).strip()
            if text:
                spans = line.get("spans", [])
                lines.append((fitz.Rect(line["bbox"]), text, spans))
    return lines


def table_rectangles(page: fitz.Page):
    rectangles = []
    for drawing in page.get_drawings():
        fill = drawing.get("fill")
        for item in drawing.get("items", []):
            if item[0] != "re":
                continue
            rect = fitz.Rect(item[1])
            if rect.width < 12 or rect.height < 8 or rect.width >= page.rect.width - 2 or rect.height >= page.rect.height - 2:
                continue
            rectangles.append((rect, fill))
    return rectangles


def grid_from_rectangles(rectangles):
    xs = unique([coordinate for rect, _ in rectangles for coordinate in (rect.x0, rect.x1)])
    ys = unique([coordinate for rect, _ in rectangles for coordinate in (rect.y0, rect.y1)])
    if len(rectangles) < 8 or len(xs) < 3 or len(ys) < 3:
        return None
    x_hits = [sum(abs(edge - point) <= 1.5 for rect, _ in rectangles for edge in (rect.x0, rect.x1)) for point in xs]
    y_hits = [sum(abs(edge - point) <= 1.5 for rect, _ in rectangles for edge in (rect.y0, rect.y1)) for point in ys]
    # A table has repeated vertical and horizontal boundaries. Paragraph
    # highlights tend to share a single left edge but have unique right edges.
    if sum(hits >= 3 for hits in x_hits) < 3 or sum(hits >= 3 for hits in y_hits) < 3:
        return None
    return xs, ys


def index_of(values: list[float], target: float) -> int:
    return min(range(len(values)), key=lambda index: abs(values[index] - target))


def set_table_sheet(sheet, page: fitz.Page, rectangles, xs, ys):
    sheet.sheet_view.showGridLines = False
    for index in range(len(xs) - 1):
        sheet.column_dimensions[get_column_letter(index + 1)].width = max(5, min(42, (xs[index + 1] - xs[index]) / 5.3))
    for index in range(len(ys) - 1):
        sheet.row_dimensions[index + 1].height = max(18, min(96, (ys[index + 1] - ys[index]) * 1.15))

    occupied = set()
    for rect, fill in rectangles:
        left, right = index_of(xs, rect.x0) + 1, index_of(xs, rect.x1)
        top, bottom = index_of(ys, rect.y0) + 1, index_of(ys, rect.y1)
        if right < left or bottom < top:
            continue
        cell = sheet.cell(top, left)
        cell.fill = PatternFill("solid", fgColor=rgb(fill))
        for row in sheet.iter_rows(min_row=top, max_row=bottom, min_col=left, max_col=right):
            for item in row:
                item.border = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)
                item.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                occupied.add((item.row, item.column))
        if right > left or bottom > top:
            try:
                sheet.merge_cells(start_row=top, start_column=left, end_row=bottom, end_column=right)
            except ValueError:
                pass

    outside_text = []
    table_box = fitz.Rect(xs[0], ys[0], xs[-1], ys[-1])
    for bbox, text, spans in text_lines(page):
        center_x, center_y = (bbox.x0 + bbox.x1) / 2, (bbox.y0 + bbox.y1) / 2
        if not table_box.contains(fitz.Point(center_x, center_y)):
            outside_text.append(text)
            continue
        column = max(1, min(len(xs) - 1, next((index + 1 for index in range(len(xs) - 1) if xs[index] <= center_x <= xs[index + 1]), 1)))
        row = max(1, min(len(ys) - 1, next((index + 1 for index in range(len(ys) - 1) if ys[index] <= center_y <= ys[index + 1]), 1)))
        target = sheet.cell(row, column)
        if isinstance(target, MergedCell):
            for merged in sheet.merged_cells.ranges:
                if target.coordinate in merged:
                    target = sheet.cell(merged.min_row, merged.min_col)
                    break
        if target.value:
            target.value = f"{target.value}\n{text}"
        else:
            target.value = text
        target.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if spans:
            size = max(span.get("size", 10) for span in spans)
            raw_color = spans[0].get("color", 0)
            target.font = Font(name=FONT_NAME, size=max(8, min(18, size)), color=f"{raw_color:06X}")

    sheet.freeze_panes = "A2"
    if outside_text:
        note_row = len(ys) + 2
        sheet.cell(note_row, 1).value = "页面标题/说明"
        sheet.cell(note_row, 1).font = Font(name=FONT_NAME, bold=True)
        sheet.cell(note_row, 2).value = "\n".join(outside_text)
        sheet.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=len(xs) - 1)
        sheet.cell(note_row, 2).alignment = Alignment(vertical="top", wrap_text=True)
        sheet.row_dimensions[note_row].height = max(24, min(120, 16 + len(sheet.cell(note_row, 2).value) / 40 * 16))


def set_text_sheet(sheet, page: fitz.Page):
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 100
    sheet["A1"] = f"第 {page.number + 1} 页"
    sheet["A1"].font = Font(name=FONT_NAME, size=15, bold=True)
    sheet.merge_cells("A1:B1")
    for index, (bbox, text, spans) in enumerate(text_lines(page), start=2):
        sheet.cell(index, 1).value = index - 1
        cell = sheet.cell(index, 2)
        cell.value = text
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if spans:
            cell.font = Font(name=FONT_NAME, size=max(9, min(16, max(span.get("size", 10) for span in spans))))
        sheet.row_dimensions[index].height = max(20, min(100, 15 + len(text) / 45 * 16))
    sheet.freeze_panes = "A2"


def sheet_name(value: str, fallback: str) -> str:
    value = re.sub(r"[\\/*?:\[\]]", "-", value).strip()
    return (value or fallback)[:31]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input_pdf")
    parser.add_argument("output_xlsx")
    args = parser.parse_args()

    pdf = fitz.open(args.input_pdf)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, page in enumerate(pdf, start=1):
        rectangles = table_rectangles(page)
        grid = grid_from_rectangles(rectangles)
        sheet = workbook.create_sheet(sheet_name(f"第{index}页", f"Page {index}"))
        if grid:
            set_table_sheet(sheet, page, rectangles, *grid)
        else:
            set_text_sheet(sheet, page)
    workbook.properties.title = pdf.metadata.get("title") or Path(args.input_pdf).stem
    workbook.save(args.output_xlsx)
    pdf.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
